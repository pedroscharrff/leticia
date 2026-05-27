"""
Inventory connector service.
Supports: manual (CRUD), rest_api, sql, webhook, csv, xlsx, google_sheets.
Each connector syncs products into the tenant's products table.
"""
from __future__ import annotations

import asyncio
import csv
import io
import json
import re
import time
from typing import Any

import httpx
import structlog

from db.postgres import get_db_conn, tenant_conn
from services import secrets as sec_svc

log = structlog.get_logger()


# ── Connector base ────────────────────────────────────────────────────────────

class InventoryConnector:
    source: str = ""

    async def fetch_products(self, config: dict, credentials: dict) -> list[dict]:
        raise NotImplementedError

    async def sync(self, tenant_id: str, schema: str, config: dict) -> dict:
        """Run a full sync and log results. If config['deactivate_missing'] is True,
        products from this source that didn't appear in the feed are set active=false."""
        start = time.monotonic()
        errors: list[str] = []
        records_in = records_upd = records_deactivated = 0
        seen_keys: list[str] = []

        try:
            credentials = await _load_credentials(tenant_id, config.get("credentials_key", ""))
            products = await self.fetch_products(config, credentials)
            records_in = len(products)

            for product in products:
                try:
                    async with tenant_conn(schema) as conn:
                        await _upsert_product(conn, product, self.source)
                        records_upd += 1
                        key = product.get("sku") or product.get("barcode") or product.get("name")
                        if key:
                            seen_keys.append(str(key))
                except Exception as exc:
                    errors.append(str(exc))

            if config.get("deactivate_missing") and seen_keys:
                try:
                    async with tenant_conn(schema) as conn:
                        records_deactivated = await _deactivate_missing(
                            conn, self.source, seen_keys
                        )
                except Exception as exc:
                    errors.append(f"deactivate_missing_failed: {exc}")
        except Exception as exc:
            errors.append(f"fetch_error: {exc}")

        duration = int((time.monotonic() - start) * 1000)
        status = "ok" if not errors else ("partial" if records_upd > 0 else "error")

        async with tenant_conn(schema) as conn:
            await conn.execute(
                """
                INSERT INTO inventory_sync_log (connector, status, records_in, records_upd, errors, duration_ms)
                VALUES ($1, $2, $3, $4, $5, $6)
                """,
                self.source, status, records_in, records_upd, json.dumps(errors), duration,
            )

        log.info("inventory.sync", tenant=tenant_id, source=self.source,
                 records_in=records_in, records_upd=records_upd,
                 records_deactivated=records_deactivated, errors=len(errors))
        # Loga cada erro individualmente para aparecer no stdout/docker logs
        for err_msg in errors:
            log.warning("inventory.sync.error_detail",
                        tenant=tenant_id, source=self.source, error=err_msg)
        return {
            "status": status,
            "records_in": records_in,
            "records_upd": records_upd,
            "records_deactivated": records_deactivated,
            "errors": errors,
        }


async def _load_credentials(tenant_id: str, key: str) -> dict:
    if not key:
        return {}
    raw = await sec_svc.get_secret(tenant_id, key)
    return json.loads(raw) if raw else {}


async def _upsert_product(conn, product: dict, source: str) -> None:
    # Limpa campos vazios/None problemáticos
    price = product.get("price")
    if isinstance(price, str):
        price = _parse_price(price)
    stock_qty = product.get("stock_qty", 0)
    if isinstance(stock_qty, str):
        stock_qty = _parse_int(stock_qty)

    await conn.execute(
        """
        INSERT INTO products (sku, name, brand, category, description, price, stock_qty, unit, barcode, source, tags, meta)
        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12)
        ON CONFLICT (sku) DO UPDATE SET
            name=EXCLUDED.name, brand=EXCLUDED.brand, category=EXCLUDED.category,
            description=EXCLUDED.description, price=EXCLUDED.price, stock_qty=EXCLUDED.stock_qty,
            unit=EXCLUDED.unit, barcode=EXCLUDED.barcode, tags=EXCLUDED.tags,
            meta=EXCLUDED.meta, source=EXCLUDED.source, active=TRUE, updated_at=NOW()
        """,
        product.get("sku"), product.get("name", ""), product.get("brand"),
        product.get("category"), product.get("description"), price,
        stock_qty or 0, product.get("unit", "un"), product.get("barcode"),
        source, product.get("tags", []) or [], json.dumps(product.get("meta", {})),
    )


async def _deactivate_missing(conn, source: str, seen_keys: list[str]) -> int:
    """Marca active=false produtos da mesma source que não vieram no feed."""
    row = await conn.fetchrow(
        """
        WITH updated AS (
            UPDATE products
            SET active = FALSE, updated_at = NOW()
            WHERE source = $1
              AND active = TRUE
              AND COALESCE(sku, barcode, name) <> ALL($2::text[])
            RETURNING id
        )
        SELECT COUNT(*) AS n FROM updated
        """,
        source, seen_keys,
    )
    return int(row["n"]) if row else 0


def _parse_price(v: Any) -> float | None:
    if v is None or v == "":
        return None
    s = str(v).strip().replace("R$", "").replace(" ", "")
    # "1.234,56" → "1234.56"
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _parse_int(v: Any) -> int:
    if v is None or v == "":
        return 0
    try:
        return int(float(str(v).strip().replace(",", ".")))
    except ValueError:
        return 0


# ── REST API Connector ────────────────────────────────────────────────────────

class RestApiConnector(InventoryConnector):
    source = "rest_api"

    async def fetch_products(self, config: dict, credentials: dict) -> list[dict]:
        base_url: str = config["base_url"]
        endpoint: str = config.get("endpoint", "/products")
        auth_type: str = config.get("auth_type", "bearer")
        mapping: dict = config.get("field_mapping", {})

        headers: dict = {}
        if auth_type == "bearer":
            headers["Authorization"] = f"Bearer {credentials.get('token','')}"
        elif auth_type == "api_key":
            headers[credentials.get("header_name", "X-Api-Key")] = credentials.get("api_key", "")
        elif auth_type == "basic":
            import base64
            cred_str = f"{credentials.get('username','')}:{credentials.get('password','')}"
            encoded = base64.b64encode(cred_str.encode()).decode()
            headers["Authorization"] = f"Basic {encoded}"

        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.get(f"{base_url}{endpoint}", headers=headers)
        resp.raise_for_status()

        data = resp.json()
        items = data if isinstance(data, list) else data.get(config.get("list_key", "data"), [])

        return [_apply_mapping(item, mapping) for item in items]


# ── SQL Connector ─────────────────────────────────────────────────────────────

class SqlConnector(InventoryConnector):
    source = "sql"

    async def fetch_products(self, config: dict, credentials: dict) -> list[dict]:
        dsn: str = credentials.get("dsn", "")
        query: str = config.get("query", "SELECT * FROM products")
        mapping: dict = config.get("field_mapping", {})

        if not dsn:
            raise ValueError("SQL connector: DSN not configured")

        try:
            import asyncpg
            conn = await asyncpg.connect(dsn, command_timeout=30, statement_timeout=30000)
            try:
                rows = await conn.fetch(query)
            finally:
                await conn.close()
        except Exception as exc:
            raise RuntimeError(f"SQL sync failed: {exc}") from exc

        return [_apply_mapping(dict(r), mapping) for r in rows]


# ── CSV Connector (import from upload) ───────────────────────────────────────

class CsvConnector(InventoryConnector):
    source = "csv"

    async def import_csv(self, tenant_id: str, schema: str, content: bytes, mapping: dict) -> dict:
        rows = _read_csv_rows(content)
        products = [_apply_mapping(r, mapping) for r in rows]
        return await _bulk_import(schema, products, self.source)

    async def fetch_products(self, config: dict, credentials: dict) -> list[dict]:
        return []  # CSV import is manual, not scheduled


# ── Excel (.xlsx) Connector ───────────────────────────────────────────────────

class XlsxConnector(InventoryConnector):
    source = "xlsx"

    async def import_xlsx(self, tenant_id: str, schema: str, content: bytes, mapping: dict) -> dict:
        rows = _read_xlsx_rows(content)
        products = [_apply_mapping(r, mapping) for r in rows]
        return await _bulk_import(schema, products, self.source)

    async def fetch_products(self, config: dict, credentials: dict) -> list[dict]:
        return []


# ── Google Sheets Connector ───────────────────────────────────────────────────

class GoogleSheetsConnector(InventoryConnector):
    """Lê uma planilha do Google Sheets em CSV.
    Estratégias (tentadas em ordem até alguma devolver CSV válido):
      1. URL "Publicar na web" (.../pub?output=csv) — recomendada, sem rate-limit
      2. /export?format=csv (precisa de 'Qualquer pessoa com o link')
      3. /gviz/tq?tqx=out:csv (fallback)
    """
    source = "google_sheets"

    async def fetch_products(self, config: dict, credentials: dict) -> list[dict]:
        url_or_id: str = config.get("sheet_url") or config.get("sheet_id") or ""
        gid: str = str(config.get("gid", "0"))
        mapping: dict = config.get("field_mapping", {})

        if not url_or_id:
            raise ValueError("Google Sheets: sheet_url ou sheet_id é obrigatório")

        candidates = _build_google_sheets_csv_candidates(url_or_id, gid)
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
            ),
            "Accept": "text/csv, */*;q=0.5",
        }

        attempt_errors: list[str] = []
        rows: list[dict] = []

        async with httpx.AsyncClient(timeout=60, follow_redirects=True) as client:
            for label, url in candidates:
                try:
                    resp = await client.get(url, headers=headers)
                except httpx.HTTPError as exc:
                    attempt_errors.append(f"{label}: {exc.__class__.__name__}: {exc}")
                    continue

                final_url = str(resp.url)
                content_type = resp.headers.get("content-type", "")
                body_preview = resp.text[:200].replace("\n", " ") if resp.text else ""

                # Tratamento .gviz: vem como "google.visualization.Query.setResponse({..."
                # — não suportamos no momento (precisaria parser JSON), então pulamos.
                # Mantemos só a checagem de HTML/erro.
                if resp.status_code != 200:
                    attempt_errors.append(
                        f"{label}: HTTP {resp.status_code} ({final_url}) preview={body_preview!r}"
                    )
                    log.warning("inventory.google_sheets.attempt_failed",
                                strategy=label, status=resp.status_code,
                                final_url=final_url, preview=body_preview)
                    continue

                looks_like_html = (
                    "text/html" in content_type.lower()
                    or body_preview.lstrip().lower().startswith(("<!doctype", "<html"))
                )
                if looks_like_html:
                    attempt_errors.append(
                        f"{label}: HTML em vez de CSV ({final_url})"
                    )
                    log.warning("inventory.google_sheets.html_response",
                                strategy=label, final_url=final_url,
                                content_type=content_type, preview=body_preview)
                    continue

                # gviz devolve JS, não CSV — detecta e pula
                if body_preview.lstrip().startswith(")]}'") or "setResponse(" in body_preview:
                    attempt_errors.append(
                        f"{label}: resposta em formato gviz/JSON, não suportada"
                    )
                    continue

                try:
                    rows = _read_csv_rows(resp.content)
                except Exception as exc:
                    attempt_errors.append(f"{label}: erro parseando CSV: {exc}")
                    continue

                if rows:
                    log.info("inventory.google_sheets.success",
                             strategy=label, rows=len(rows))
                    break  # sucesso!
                attempt_errors.append(f"{label}: 0 linhas no CSV")

        # Se o usuário não definiu mapping, aplica auto-detect pelos cabeçalhos.
        # Assim quem usa o atalho "Conectar Google Sheets" sem passar pelo preview
        # também acerta os campos quando as colunas seguem nomes comuns (PT-BR).
        if rows and not mapping:
            mapping = suggest_mapping(list(rows[0].keys()))
            log.info("inventory.google_sheets.auto_mapping",
                     mapping=mapping, headers=list(rows[0].keys()))

        if not rows:
            raise RuntimeError(
                "Não consegui ler a planilha por nenhuma estratégia disponível. "
                "Solução recomendada: no Google Sheets, vá em "
                "Arquivo → Compartilhar → Publicar na web → escolha a aba e o formato CSV "
                "→ clique Publicar → copie a URL gerada (termina em /pub?output=csv) "
                "e cole aqui no lugar da URL atual. "
                f"Tentativas: {' | '.join(attempt_errors)}"
            )

        return [_apply_mapping(r, mapping) for r in rows]


_SHEET_ID_RE = re.compile(r"/spreadsheets/d/(?:e/)?([a-zA-Z0-9-_]+)")
_PUB_URL_RE = re.compile(r"/spreadsheets/d/e/[a-zA-Z0-9-_]+/pub", re.IGNORECASE)


def _build_google_sheets_csv_candidates(url_or_id: str, gid: str) -> list[tuple[str, str]]:
    """Devolve uma lista (label, url) de estratégias a tentar em ordem.

    Se o usuário colou uma URL de "Publicar na web" (/pub?output=csv), ela vem
    como prioridade absoluta — esse endpoint não tem rate-limit por IP.
    Caso contrário tenta /export?format=csv e /gviz/tq?tqx=out:csv.
    """
    url = url_or_id.strip()

    # Caso 1: usuário colou direto uma URL de "Publicar na web"
    if _PUB_URL_RE.search(url):
        u = url
        if "output=" not in u:
            sep = "&" if "?" in u else "?"
            u = f"{u}{sep}output=csv"
        if gid and "gid=" not in u:
            u = f"{u}&gid={gid}"
        return [("publish_to_web", u)]

    # Caso 2: URL "normal" (/edit?usp=sharing) ou só o ID
    m = _SHEET_ID_RE.search(url)
    sheet_id = m.group(1) if m else url
    return [
        ("export_csv", f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"),
        ("gviz",       f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&gid={gid}"),
    ]


# ── Tabular readers + bulk import shared ──────────────────────────────────────

def _read_csv_rows(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig", errors="replace")
    # Tenta detectar delimitador (vírgula, ponto-e-vírgula, tab)
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    return [dict(r) for r in reader if any((v or "").strip() for v in r.values())]


def _read_xlsx_rows(content: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl não instalado — adicione à requirements.txt") from exc

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        return []
    rows: list[dict] = []
    for r in rows_iter:
        if r is None:
            continue
        # ignora linhas totalmente vazias
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        record = {}
        for i, col in enumerate(header):
            if not col:
                continue
            val = r[i] if i < len(r) else None
            record[col] = val
        rows.append(record)
    wb.close()
    return rows


async def _bulk_import(schema: str, products: list[dict], source: str) -> dict:
    errors: list[str] = []
    records_upd = 0
    for product in products:
        if not product.get("name") and not product.get("sku"):
            errors.append("registro sem name/sku ignorado")
            continue
        try:
            async with tenant_conn(schema) as conn:
                await _upsert_product(conn, product, source)
                records_upd += 1
        except Exception as exc:
            errors.append(str(exc))
    return {"records_in": len(products), "records_upd": records_upd, "errors": errors}


# ── Mapping helper + auto-suggest ─────────────────────────────────────────────

def _apply_mapping(item: dict, mapping: dict) -> dict:
    """Translate external field names to internal schema.
    mapping = {internal_field: external_field}. If empty, returns item as-is
    (assumes external columns already match internal names)."""
    if not mapping:
        return item
    return {internal: item.get(external) for internal, external in mapping.items() if external in item}


# Regex de heurística para sugerir mapping a partir de nomes de colunas comuns no Brasil
_FIELD_HINTS: dict[str, list[str]] = {
    "sku":             [r"^sku$", r"^codigo$", r"^cod$", r"cod[._-]?prod", r"id[._-]?prod"],
    "barcode":         [r"barcode", r"cod[._-]?barra", r"^ean$", r"^gtin$"],
    "name":            [r"^nome$", r"^name$", r"descricao[._-]?produto", r"^produto$", r"desc[._-]?prod", r"nm[._-]?prod"],
    "brand":           [r"^marca$", r"^brand$"],
    "category":        [r"categoria", r"^category$", r"^grupo$", r"departamento"],
    "description":     [r"^descricao$", r"^description$", r"observac", r"dosagem", r"formato"],
    # price: pega "preco", "preco_promocional", "vl_venda", "valor_venda", "preco_de_venda", etc.
    "price":           [r"preco", r"^price$", r"vl[._-]?venda", r"valor", r"vlr"],
    "stock_qty":       [r"^estoque$", r"qtd[._-]?estoque", r"^qtde$", r"^quantidade$", r"^saldo$", r"^stock$"],
    "unit":            [r"^unidade$", r"^unid$", r"^un$", r"^unit$"],
    "principio_ativo": [r"principio[._-]?ativo", r"active[._-]?ingredient"],
    "fabricante":      [r"fabricante", r"laboratorio", r"^lab$"],
}


def suggest_mapping(headers: list[str]) -> dict[str, str]:
    """Dado os cabeçalhos de uma planilha, sugere um mapping internal→external."""
    suggestion: dict[str, str] = {}
    normalized = [(h, _normalize(h)) for h in headers if h]
    for internal, patterns in _FIELD_HINTS.items():
        for raw, norm in normalized:
            if any(re.search(p, norm) for p in patterns):
                suggestion[internal] = raw
                break
    return suggestion


def _normalize(s: str) -> str:
    import unicodedata
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().lower()
    return re.sub(r"[\s]+", "_", s.strip())


# ── Preview helper (used before import) ───────────────────────────────────────

def preview_tabular(content: bytes, filename: str, sample_rows: int = 5) -> dict:
    """Lê CSV ou XLSX, devolve cabeçalho + amostra + mapping sugerido."""
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        rows = _read_xlsx_rows(content)
    else:
        rows = _read_csv_rows(content)

    if not rows:
        return {"headers": [], "rows": [], "suggested_mapping": {}, "total_rows": 0}

    headers = list(rows[0].keys())
    sample = rows[: max(1, sample_rows)]
    # converte valores para str/None para serialização JSON segura
    sample_safe = [
        {k: (v if v is None or isinstance(v, (str, int, float, bool)) else str(v)) for k, v in r.items()}
        for r in sample
    ]
    return {
        "headers": headers,
        "rows": sample_safe,
        "suggested_mapping": suggest_mapping(headers),
        "total_rows": len(rows),
    }


# ── PDV Templates (presets de mapping para sistemas brasileiros) ─────────────

PDV_TEMPLATES: list[dict] = [
    {
        "id": "generico_csv_br",
        "label": "CSV genérico (PT-BR)",
        "description": "Planilha com colunas em português: Nome, Preco, Estoque, etc.",
        "field_mapping": {
            "name": "Nome", "sku": "SKU", "barcode": "Codigo de Barras",
            "category": "Categoria", "brand": "Marca", "price": "Preco",
            "stock_qty": "Estoque", "unit": "Unidade",
            "principio_ativo": "Principio Ativo", "fabricante": "Fabricante",
        },
    },
    {
        "id": "trier",
        "label": "Trier Farmácia",
        "description": "Exportação padrão do PDV Trier (cod_prod, nm_prod, vl_venda...)",
        "field_mapping": {
            "sku": "cod_prod", "name": "nm_prod", "barcode": "cod_barra",
            "price": "vl_venda", "stock_qty": "qtd_estoque",
            "fabricante": "lab", "principio_ativo": "principio_ativo",
        },
    },
    {
        "id": "vivver",
        "label": "Vivver",
        "description": "Layout Vivver (codigo, descricao, preco_venda...)",
        "field_mapping": {
            "sku": "codigo", "name": "descricao", "barcode": "ean",
            "price": "preco_venda", "stock_qty": "saldo_estoque",
            "category": "grupo", "fabricante": "fabricante",
        },
    },
    {
        "id": "linx_big",
        "label": "Linx Big Farma",
        "description": "Layout Linx Big Farma (CODIGO, DESCRICAO, PRECO...)",
        "field_mapping": {
            "sku": "CODIGO", "name": "DESCRICAO", "barcode": "CODBARRA",
            "price": "PRECO", "stock_qty": "ESTOQUE",
            "category": "DEPARTAMENTO", "fabricante": "LABORATORIO",
        },
    },
    {
        "id": "epharma",
        "label": "ePharma",
        "description": "Layout ePharma (cod_ean, descricao_produto, vlr_unit...)",
        "field_mapping": {
            "barcode": "cod_ean", "name": "descricao_produto",
            "price": "vlr_unit", "stock_qty": "qtd_disponivel",
            "fabricante": "laboratorio",
        },
    },
    {
        "id": "sysmo",
        "label": "Sysmo",
        "description": "Layout Sysmo (CD_PRODUTO, NM_PRODUTO, VL_VENDA...)",
        "field_mapping": {
            "sku": "CD_PRODUTO", "name": "NM_PRODUTO", "barcode": "CD_BARRAS",
            "price": "VL_VENDA", "stock_qty": "QT_ESTOQUE",
            "category": "GRUPO", "fabricante": "FABRICANTE",
        },
    },
]


def get_template(template_id: str) -> dict | None:
    return next((t for t in PDV_TEMPLATES if t["id"] == template_id), None)


CONNECTOR_REGISTRY: dict[str, type[InventoryConnector]] = {
    "rest_api": RestApiConnector,
    "sql": SqlConnector,
    "csv": CsvConnector,
    "xlsx": XlsxConnector,
    "google_sheets": GoogleSheetsConnector,
}
